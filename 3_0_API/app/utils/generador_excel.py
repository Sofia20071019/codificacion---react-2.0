"""
ARCHIVO: generador_excel.py
PROPOSITO: Generador de archivos Excel (.xlsx) para los reportes del sistema
           Kimuka. Construye un workbook con:
             - Una hoja "Resumen" con el nombre del reporte, fecha de generacion,
               usuario, filtros utilizados y estadisticas clave.
             - Una o mas hojas de datos con encabezados formateados, ancho de
               columnas, autofiltro, congelacion de encabezados, formatos de
               fecha/numero y filas de totales.
           Utiliza la libreria openpyxl y no almacena nada en la base de datos:
           el archivo se genera en memoria (BytesIO) y se devuelve como descarga.
"""

# Importacion de io para generar el archivo en memoria sin tocar el disco
import io

# Importacion de unicodedata y re para construir nombres de archivo seguros
import re
import unicodedata

# Importacion de datetime para la fecha de generacion del reporte
from datetime import date, datetime

# Importacion de las clases de openpyxl para crear el workbook y aplicar estilos
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------------------
# ESTILOS REUTILIZABLES
# ------------------------------------------------------------------------------

# Fuente del titulo principal (negrita, azul oscuro corporativo)
TITULO_FONT = Font(bold=True, size=14, color="1F4E78")

# Fuente de las etiquetas de metadatos en la hoja Resumen
ETIQUETA_FONT = Font(bold=True, size=10, color="404040")

# Fuente de los valores de metadatos en la hoja Resumen
VALOR_FONT = Font(size=10)

# Relleno azul oscuro para las celdas de encabezado de tablas
ENCABEZADO_FILL = PatternFill("solid", fgColor="1F4E78")

# Fuente blanca en negrita para los encabezados de tablas
ENCABEZADO_FONT = Font(bold=True, color="FFFFFF", size=10)

# Relleno gris azulado para las filas de totales
TOTAL_FILL = PatternFill("solid", fgColor="DCE6F1")

# Fuente en negrita para las filas de totales
TOTAL_FONT = Font(bold=True, size=10, color="1F4E78")

# Borde fino para todas las celdas de las tablas
BORDE_FINO = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Alineacion centrada y vertical para encabezados
ALINEACION_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ------------------------------------------------------------------------------
# FUNCIONES AUXILIARES
# ------------------------------------------------------------------------------


def normalizar_texto(texto):
    """Elimina acentos y caracteres especiales de una cadena."""
    if not texto:
        return "reporte"
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    texto = re.sub(r"[^A-Za-z0-9 _-]", "", texto).strip()
    return re.sub(r"\s+", "_", texto) or "reporte"


def nombre_archivo_excel(nombre_reporte):
    """
    Construye un nombre de archivo seguro a partir del nombre del reporte
    y la fecha actual. Ejemplo: Reporte_Materias_Primas_2026-08-09.xlsx.
    """
    base = normalizar_texto(nombre_reporte)
    hoy = date.today().strftime("%Y-%m-%d")
    return f"{base}_{hoy}.xlsx"


def _coercion_fecha(valor):
    """
    Convierte cadenas en formato 'YYYY-MM-DD' (o prefijo de una fecha)
    a objetos date para que Excel aplique formatos de fecha correctamente.
    Si el valor no es una fecha reconocible, se devuelve sin cambios.
    """
    if isinstance(valor, (date, datetime)):
        return valor
    if isinstance(valor, str):
        try:
            return datetime.strptime(valor[:10], "%Y-%m-%d").date()
        except (ValueError, IndexError):
            return valor
    return valor


def _escribir_fila(ws, fila_num, fila, formatos=None):
    """Escribe una fila aplicando conversion de fechas, bordes y formatos."""
    for col, valor in enumerate(fila, start=1):
        celda = ws.cell(row=fila_num, column=col)
        celda.value = _coercion_fecha(valor)
        celda.border = BORDE_FINO
        if formatos and col in formatos:
            celda.number_format = formatos[col]


def _escribir_encabezados(ws, fila_num, encabezados):
    """Escribe los encabezados de una tabla con estilo de encabezado."""
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_num, column=col)
        celda.value = texto
        celda.font = ENCABEZADO_FONT
        celda.fill = ENCABEZADO_FILL
        celda.alignment = ALINEACION_CENTRO
        celda.border = BORDE_FINO


def _aplicar_anchos(ws, anchos):
    """Configura el ancho de cada columna de la hoja."""
    for idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho


# ------------------------------------------------------------------------------
# CONSTRUCTOR PRINCIPAL
# ------------------------------------------------------------------------------


def crear_excel_reporte(nombre_reporte, usuario, filtros, estadisticas, secciones):
    """
    Construye el workbook de un reporte completo.

    Args:
        nombre_reporte (str): Nombre del reporte (ej: 'Reporte de Materias Primas').
        usuario (str): Nombre del usuario que genera el reporte.
        filtros (list): Lista de tuplas (etiqueta, valor) con los filtros usados.
        estadisticas (list): Lista de tuplas (etiqueta, valor) con estadisticas clave.
        secciones (list): Lista de dicts con la estructura de cada hoja de datos:
            {
                "hoja": nombre de la hoja (max 31 caracteres),
                "titulo": titulo descriptivo de la seccion,
                "encabezados": [nombre de columnas],
                "filas": [list de valores por fila],
                "ancho_columnas": [ancho por columna] (opcional),
                "formatos": {col_idx: number_format} (opcional, 1-based),
                "totales": [valores de la fila de total] (opcional)
            }

    Returns:
        io.BytesIO: Archivo .xlsx en memoria listo para enviarse al cliente.
    """
    wb = Workbook()

    # ------------------------------------------------------------------
    # HOJA: RESUMEN (metadatos + estadisticas + filtros)
    # ------------------------------------------------------------------
    resumen = wb.active
    resumen.title = "Resumen"

    fila = 1
    resumen.cell(fila, 1, f"SISTEMA KIMUKA - {nombre_reporte.upper()}").font = TITULO_FONT
    fila += 2

    # Metadatos del reporte
    resumen.cell(fila, 1, "Fecha de generacion:").font = ETIQUETA_FONT
    resumen.cell(fila, 2, datetime.now().strftime("%d/%m/%Y %H:%M")).font = VALOR_FONT
    fila += 1
    resumen.cell(fila, 1, "Usuario:").font = ETIQUETA_FONT
    resumen.cell(fila, 2, usuario or "-").font = VALOR_FONT
    fila += 2

    # Filtros utilizados
    resumen.cell(fila, 1, "FILTROS UTILIZADOS").font = ETIQUETA_FONT
    fila += 1
    for etiqueta, valor in filtros:
        resumen.cell(fila, 1, f"- {etiqueta}:").font = VALOR_FONT
        resumen.cell(fila, 2, valor or "Todos").font = VALOR_FONT
        fila += 1
    fila += 1

    # Estadisticas clave
    resumen.cell(fila, 1, "ESTADISTICAS").font = ETIQUETA_FONT
    fila += 1
    for etiqueta, valor in estadisticas:
        resumen.cell(fila, 1, f"- {etiqueta}:").font = VALOR_FONT
        resumen.cell(fila, 2, valor).font = VALOR_FONT
        fila += 1

    # Hoja de resumen: ancho de columnas
    resumen.column_dimensions["A"].width = 35
    resumen.column_dimensions["B"].width = 40

    # ------------------------------------------------------------------
    # HOJAS DE DATOS (una por seccion)
    # ------------------------------------------------------------------
    for seccion in secciones:
        hoja = wb.create_sheet(title=seccion["hoja"])
        encabezados = seccion["encabezados"]
        filas = seccion.get("filas", [])
        formatos = seccion.get("formatos", {})
        anchos = seccion.get("ancho_columnas", [20] * len(encabezados))

        # Titulo de la seccion en la primera fila
        hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
        celda_titulo = hoja.cell(1, 1, seccion["titulo"])
        celda_titulo.font = TITULO_FONT

        # Encabezados en la fila 3
        fila_encabezados = 3
        _escribir_encabezados(hoja, fila_encabezados, encabezados)

        # Datos a partir de la fila 4
        fila_datos = fila_encabezados + 1
        filas_datos = list(filas) if filas else [["Sin datos para el filtro seleccionado."] + [None] * (len(encabezados) - 1)]

        for idx, fila_valores in enumerate(filas_datos):
            _escribir_fila(hoja, fila_datos + idx, fila_valores, formatos)

        # Fila de totales (opcional)
        totales = seccion.get("totales")
        fila_totales = fila_datos + len(filas_datos)
        if totales:
            _escribir_fila(hoja, fila_totales, totales, formatos)
            for col in range(1, len(encabezados) + 1):
                celda = hoja.cell(fila_totales, col)
                celda.font = TOTAL_FONT
                celda.fill = TOTAL_FILL
            fila_totales += 1

        # Autofiltro sobre la fila de encabezados y todos los datos
        if filas_datos:
            ultima_fila = fila_totales - 1 if totales else fila_datos + len(filas_datos) - 1
            if ultima_fila >= fila_encabezados:
                hoja.auto_filter.ref = f"A{fila_encabezados}:{get_column_letter(len(encabezados))}{ultima_fila}"

        # Congelar panes: se mantienen visibles titulo y encabezados
        hoja.freeze_panes = f"A{fila_encabezados + 1}"

        # Ancho de columnas
        _aplicar_anchos(hoja, anchos)

    # Escribir el workbook en un buffer en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
